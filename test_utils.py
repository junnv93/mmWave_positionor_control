# test_utils.py
import os
import time
import re
import pyvisa
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import openpyxl
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import QFileDialog
from logger_config import get_logger  

logger = get_logger()


def is_excel_file(file_path):
    return re.search(r'\.xlsx$', file_path, re.IGNORECASE) is not None



def load_excel_file():
    try:
        file_path, _ = QFileDialog.getOpenFileName(None, "Select Excel file", "", "Excel files (*.xlsx)")
        if not file_path:
            logger.warning('Excel file selection was cancelled.')
            return None
        if not is_excel_file(file_path):
            logger.error('Selected file is not an Excel file.')
            return None
        return file_path
    except Exception as e:
        logger.error(f"Error loading Excel file: {e}")
        return None


def read_excel_sheet(file_path, sheet_name):
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name, header=0)
    except Exception as e:
        print(f"Error reading sheet {sheet_name}: {e}")
        return None

def read_gpib_addresses(file_path):
    try:
        # Assuming 'Chamber Config' sheet has consistent format with GPIB addresses at specific locations
        config_df = pd.read_excel(file_path, sheet_name='Chamber Config', usecols="B", header=None, nrows=3)
        analyzer_gpib = config_df.iloc[1, 0]  # Read only the required cell directly
        bt_tester_gpib = config_df.iloc[2, 0]

        # Optional: Validate GPIB addresses to ensure they are integers
        analyzer_gpib = str(analyzer_gpib)
        bt_tester_gpib = str(bt_tester_gpib)

        return {
            'Analyzer GPIB': analyzer_gpib,
            'BT Tester GPIB': bt_tester_gpib,
        }
    except ValueError as e:
        logger.error(f"Non-integer GPIB address found: {e}")
        return {}
    except Exception as e:
        logger.error(f"Error reading GPIB addresses from Chamber Config: {e}")
        return {}


def is_valid_match(test_params, row, keys):
    return all(test_params[key] == row[key] for key in keys)

def match_frequency_table_center_frequency(test_params, frequency_table_df):
    try:
        for index, row in frequency_table_df.iterrows():
            bands = str(row['Band']).split(';')
            if test_params['Band'] not in bands:
                continue
            if is_valid_match(test_params, row, ['Technology', 'Bandwidth', 'Channel']):
                return row['Center Frequency']
        return None
    except Exception as e:
        logger.error(f"Error in matching frequency table: {e}")
        return None



def match_analyzer_settings(test_params, analyzer_settings_df, param_mapping):
    for index, row in analyzer_settings_df.iterrows():
        match_found = True

        for key, value in param_mapping.items():
            if key in ['Channel', 'Number of Resource Blocks'] and pd.isna(row.get(value)):
                continue  # Channel 정보가 없는 경우 비교하지 않음

            if key in ['Channel', 'Band', 'Technology','Number of Carriers']:
                setting_values = str(row.get(value, '')).split(';')
                param_value = test_params.get(key)

                # Channel 값이 숫자인 경우 정수로 변환하여 문자열로 처리
                if key == 'Channel':
                    if isinstance(param_value, (int, float)):
                        param_value_str = str(int(param_value))
                    else:
                        param_value_str = str(param_value)
                else:
                    param_value_str = str(param_value)

                if param_value_str not in setting_values:
                    match_found = False
                    logger.info(f"Matching failed for key '{key}', Test Param: '{param_value_str}', Row Values: {setting_values}")
                    break
            else:
                if str(test_params.get(key)) != str(row.get(value)):
                    match_found = False
                    logger.info(f"Matching failed for key '{key}', Test Param: '{test_params.get(key)}', Row Value: '{row.get(value)}'")
                    break

        if match_found:
            logger.info(f"Matching successful for row index {index}: {row.to_dict()}")
            return {col: row[col] for col in analyzer_settings_df.columns[5:]}
        # else:
        #     logger.info(f"No match found for test parameters: {test_params}")

    return None

def save_updated_test_plan(file_path, test_plan_df, sheet_name, column_to_update):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # 업데이트할 컬럼이 첫 번째라고 가정
        for index, row in test_plan_df.iterrows():
            if str(row.iloc[0]).startswith('#'):
                # OpenPyXL은 1부터 인덱싱, 행과 열 인덱스에 1을 더함
                sheet.cell(row=index + 2, column=column_to_update).value = row.iloc[0]

        workbook.save(file_path)
    except Exception as e:
        logger.error(f"테스트 플랜 업데이트 중 오류 발생: {e}")

def open_instrument(address, resource_manager):
    try:
        instrument = resource_manager.open_resource(address)
        instrument.timeout = 5000  # Set timeout to 5 seconds
        return instrument
    except pyvisa.VisaIOError as e:
        logger.error(f"Error opening connection to the spectrum analyzer: {e}")
        return None
    

def read_save_data(file_path):
    """
    Excel 파일의 'Save Data' 시트에서 데이터를 읽습니다.
    """
    save_data_df = read_excel_sheet(file_path, 'Save Data')
    if save_data_df is None:
        logger.error("Error: 'Save Data' 시트를 로드할 수 없습니다.")
        return None

    # 필요한 정보 추출
    model_number = save_data_df.iloc[0, 1]
    sample_no = save_data_df.iloc[1, 1]
    user_id = save_data_df.iloc[2, 1]
    base_folder_path = save_data_df.iloc[3, 1] # 데이터 저장을 위한 기본 폴더 경로

    return {
        'Model Number': model_number,
        'Sample No': sample_no,
        'User ID': user_id,
        'Base Folder Path': base_folder_path
    }     

def generate_file_name(test_params, additional_suffix=''):
    file_name_parts = [
        str(test_params.get('Test', '')),
        str(test_params.get('Band', '')),
        str(test_params.get('Antenna', '')),
        str(test_params.get('Mode', '')),
        str(test_params.get('Number of Carriers', '')),
        str(test_params.get('Modulation', '')),
        str(test_params.get('Bandwidth', '')),
        str(test_params.get('Channel', '')) + additional_suffix,
        str(test_params.get('Beam ID', '')),
        str(test_params.get('Resource Block Start', '')),
        str(test_params.get('Number of Resource Blocks', ''))
    ]

    file_name_parts = [part if part != 'nan' else '' for part in file_name_parts]

    return '_'.join(filter(None, file_name_parts))



def update_data_in_excel(file_path, sheet_name, match_columns, test_params, result1, result2=None):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    header = [cell.value for cell in sheet[1]]  # 헤더 행 읽기

    # 결과 컬럼 이름 설정
    result_column1 = 'On Time' if test_params['Test'] == 'Duty' else 'Result1'
    result_column2 = 'Period' if test_params['Test'] == 'Duty' else ('Result2(optional)' if result2 is not None else None)

    # 매칭 컬럼의 인덱스 찾기
    match_indices = [header.index(col) for col in match_columns if col in header]
    result_index1 = header.index(result_column1) if result_column1 in header else None
    result_index2 = header.index(result_column2) if result_column2 and result_column2 in header else None

    update_successful = False  # Flag to track success

    for row in sheet.iter_rows(min_row=2):
        row_values = [cell.value for cell in row]
        # logger.info(f"Checking row: {row_values}")

        if all(pd.isna(row[idx].value) if pd.isna(test_params[col]) else row[idx].value == test_params[col] 
               for idx, col in zip(match_indices, match_columns)):
            logger.info(f"Match found: {row_values}")
            if result_index1 is not None and result1 is not None:
                row[result_index1].value = result1
                logger.info(f"Updated Result1: {result1}")

            if result_index2 is not None and result2 is not None:
                row[result_index2].value = result2
                logger.info(f"Updated Result2: {result2}")

            update_successful = True  # Update successful flag set
            break
        else:
            # 실패한 매칭에 대한 정보 로그에 기록
            failed_matches = [(col, test_params[col], row[idx].value) for idx, col in zip(match_indices, match_columns) if not (pd.isna(row[idx].value) if pd.isna(test_params[col]) else row[idx].value == test_params[col])]
            # logger.info(f"Match failed for row: {row_values}. Failed matches: {failed_matches}")

    workbook.save(file_path)
    workbook.close()
    return update_successful

def record_notification(file_path, notification_type, message):
    workbook = openpyxl.load_workbook(file_path)
    sheet_name = 'Notification'
    
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]

    # Find the first empty row in the sheet
    max_row = sheet.max_row + 1 if sheet.max_row > 1 else 1
    
    # Insert the notification
    sheet[f"A{max_row}"] = notification_type
    sheet[f"B{max_row}"] = message

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width


    workbook.save(file_path)
    workbook.close()


def capture_and_save_screen(test_params, save_data, file_path, additional_suffix=''):
    rm = pyvisa.ResourceManager()
    device = None
    try:
        # Chamber Config 시트에서 GPIB 주소 읽기
        gpib_addresses = read_gpib_addresses(file_path)
        analyzer_gpib = gpib_addresses.get('Analyzer GPIB', '18')  # 기본값 '18'

        # GPIB 주소를 사용하여 장비에 연결
        device = open_instrument(f'TCPIP0::{analyzer_gpib}::inst0::INSTR', rm)
        
        if device is None:
            logger.error("Failed to connect to the instrument.")
            raise Exception("Failed to connect to the device.")
        
        # Create file path for local storage
        base_path = save_data.get('Base Folder Path')
        if not base_path:
            logger.error("Base Folder Path is missing in save data. Ensure that it's correctly loaded and passed to this function.")
            raise ValueError("Base folder path is missing in save data.")
        
        model_number = str(save_data['Model Number'])  # 문자열로 변환
        technology = str(test_params['Technology'])  # 문자열로 변환
        test = str(test_params['Test'])  
        band = str(test_params['Band'])  # 문자열로 변환
        antenna = str(test_params['Antenna'])  # 문자열로 변환
        mode = str(test_params['Mode'])  # 문자열로 변환
        numberofcarriers = str(test_params['Number of Carriers'])  # 문자열로 변환
        modulation = str(test_params['Modulation'])  # 문자열로 변환    
        bandwidth = str(test_params['Bandwidth'])  # 문자열로 변환
        channel = str(test_params['Channel'])  # 문자열로 변환
        beamid = str(test_params['Beam ID'])  # 문자열로 변환

        directory_path = os.path.join(base_path, model_number, technology, test, band, antenna, mode, numberofcarriers, modulation, bandwidth, channel, beamid)
        if not os.path.exists(directory_path):
            os.makedirs(directory_path)

        file_name = generate_file_name(test_params, additional_suffix)
        device_path = f"C:\\Temp\\{file_name}.png"
        local_path = os.path.join(directory_path, f"{file_name}.png")

        # Configure device to save the screenshot
        device.query('*OPC?')
        device.write('HCOP:DEV:COL ON')
        device.write('HCOP:DEV:LANG PNG')
        device.write('HCOP:DEST "MMEM"')
        device.write(f'MMEM:NAME "{device_path}"')

        # Trigger screenshot capture
        device.write('HCOP:IMM')

        # Retrieve the screenshot
        query = f'MMEM:DATA? "{device_path}"'
        file_data = device.query_binary_values(query, datatype='B')
        with open(local_path, "wb") as file:
            file.write(bytearray(file_data))
        
        logger.info(f"Screenshot saved to {local_path}.")
    except Exception as e:
        logger.error(f"Error during screen capture: {e}, at {local_path}")
    finally:
        if device:
            device.close()
